import re
import json
import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext, ttk

import openpyxl


def parse_work_content(content):
    """
    解析工作内容，提取标题、目的、做法/过程、结果
    """
    if not content:
        return {
            "title": "",
            "purpose": "",
            "process": "",
            "result": ""
        }
    
    content_str = str(content).strip()
    lines = content_str.split('\n')
    
    title = ""
    purpose = ""
    process = ""
    result = ""
    
    current_section = "title"
    
    for line in lines:
        line = line.strip()
        if not line:
            continue
        
        # 检查是否是新的部分
        if line.startswith("目的：") or line.startswith("目的:"):
            current_section = "purpose"
            purpose = line.replace("目的：", "").replace("目的:", "").strip()
        elif line.startswith("做法：") or line.startswith("做法:") or line.startswith("过程：") or line.startswith("过程:"):
            current_section = "process"
            process_text = line.replace("做法：", "").replace("做法:", "").replace("过程：", "").replace("过程:", "").strip()
            if process_text:
                process = process_text
        elif line.startswith("结果：") or line.startswith("结果:"):
            current_section = "result"
            result = line.replace("结果：", "").replace("结果:", "").strip()
        else:
            # 根据当前部分添加内容
            if current_section == "title":
                if title:
                    title += " " + line
                else:
                    title = line
            elif current_section == "purpose":
                if purpose:
                    purpose += " " + line
                else:
                    purpose = line
            elif current_section == "process":
                if process:
                    process += " " + line
                else:
                    process = line
            elif current_section == "result":
                if result:
                    result += " " + line
                else:
                    result = line
    
    return {
        "title": title,
        "purpose": purpose,
        "process": process,
        "result": result
    }


def extract_date_range(date_string):
    """
    从日期字符串中提取开始日期和结束日期
    例如: "工作周报        日期：2024年12月02日-2024年12月06日（第48周）" 
    """
    if not date_string:
        return None, None
    
    # 转换为字符串并清理
    date_str = str(date_string).strip()
    
    # 匹配格式: 2024年12月02日-2024年12月06日 或 2024年12月02日~2024年12月06日
    match = re.search(r'(\d{4}年\d{1,2}月\d{1,2}日)\s*[-~]\s*(\d{4}年\d{1,2}月\d{1,2}日)', date_str)
    if match:
        return match.group(1), match.group(2)
    
    return None, None


def read_weekly_report(file_path):
    """
    读取单个周报Excel文件并提取信息
    返回包含日期和工作内容的字典
    """
    result = {
        "start_date": None,
        "end_date": None,
        "work_items": []
    }

    try:
        # 打开Excel文件
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        sheet = workbook.active
        
        # 1. 提取日期（第1行）
        date_cell = sheet.cell(row=1, column=1).value
        if date_cell:
            start_date, end_date = extract_date_range(str(date_cell))
            result["start_date"] = start_date
            result["end_date"] = end_date

        # 2. 提取主要工作内容（从第5行开始）

        row_num = 5
        while True:
            # 读取A列的序号
            serial_number = sheet.cell(row=row_num, column=1).value
            
            # 如果A列没有序号，停止读取
            if not serial_number:
                break
            
            # 转换为字符串检查
            serial_str = str(serial_number).strip()
            
            # 检查序号是否包含"下周工作计划"等停止词
            if "下周" in serial_str or "工作计划" in serial_str:
                break
            
            # 尝试转换为数字，如果不是数字也停止
            try:
                int(serial_str)
            except (ValueError, TypeError):
                # 不是数字序号，可能是其他内容，停止读取
                break
            
            # 读取B到F列，找到有内容的那一列
            work_content = None
            for col in range(2, 7):  # B列(2)到F列(6)
                cell_value = sheet.cell(row=row_num, column=col).value
                if cell_value and str(cell_value).strip():
                    work_content = cell_value
                    break
            
            # 如果没有找到工作内容，跳过这行
            if not work_content:
                row_num += 1
                continue
            
            # 检查内容中是否包含"下周工作计划"
            if "下周工作计划" in str(work_content):
                break
            
            # 解析工作内容
            parsed_content = parse_work_content(work_content)

            result["work_items"].append(parsed_content)

            row_num += 1

            # 安全检查：避免无限循环
            if row_num > 1000:
                break

        workbook.close()

    except Exception as e:
        raise RuntimeError(f"读取文件时出错: {e}")

    return result


def main():
    """
    主函数：提供窗口选择文件并展示解析结果
    """
    root = tk.Tk()
    root.title("周报解析器")
    root.geometry("900x650")
    root.minsize(820, 600)
    root.configure(bg="#f7f7fb")

    style = ttk.Style()
    try:
        style.theme_use("clam")
    except tk.TclError:
        pass
    style.configure("TFrame", background="#f7f7fb")
    style.configure("TLabel", background="#f7f7fb", font=("PingFang SC", 12))
    style.configure("Header.TLabel", font=("PingFang SC", 18, "bold"), foreground="#2d3a4a")
    style.configure("Accent.TButton", font=("PingFang SC", 12, "bold"), padding=(16, 10))
    style.configure("TButton", font=("PingFang SC", 12), padding=(12, 8))
    style.configure("TLabelframe", background="#f7f7fb", font=("PingFang SC", 12))
    style.configure("TLabelframe.Label", background="#f7f7fb", font=("PingFang SC", 12, "bold"))

    container = ttk.Frame(root, padding=20, style="TFrame")
    container.pack(fill=tk.BOTH, expand=True)

    header = ttk.Frame(container, style="TFrame")
    header.pack(fill=tk.X, pady=(0, 12))

    title_label = ttk.Label(header, text="周报解析器", style="Header.TLabel")
    title_label.pack(anchor="w", pady=(0, 4))

    description = ttk.Label(
        header,
        text="选择周报Excel文件后，将解析工作内容并以结构化的 JSON 数组展示。",
        wraplength=760,
        foreground="#4a5568",
    )
    description.pack(anchor="w")

    button_bar = ttk.Frame(container, style="TFrame")
    button_bar.pack(fill=tk.X, pady=(0, 12))

    output_area = ttk.LabelFrame(container, text="解析结果", padding=12)
    output_area.pack(fill=tk.BOTH, expand=True)

    output_text = scrolledtext.ScrolledText(
        output_area,
        wrap=tk.WORD,
        height=25,
        font=("JetBrains Mono", 11),
        background="#ffffff",
        relief=tk.FLAT,
        borderwidth=0,
    )
    output_text.pack(fill=tk.BOTH, expand=True)

    status_label = ttk.Label(container, text="尚未选择文件", foreground="#6b7280")
    status_label.pack(anchor="w", pady=(8, 0))

    def select_files():
        file_paths = filedialog.askopenfilenames(
            title="选择周报文件",
            filetypes=[("Excel 文件", "*.xlsx *.xls"), ("所有文件", "*.*")]
        )

        if not file_paths:
            status_label.config(text="尚未选择文件")
            return

        status_label.config(text=f"已选择 {len(file_paths)} 个文件，正在解析…")
        all_results = []
        for file_path in file_paths:
            try:
                file_result = read_weekly_report(file_path)
                all_results.append(file_result)
            except Exception as e:
                messagebox.showerror("读取失败", str(e))

        if all_results:
            json_output = json.dumps(all_results, ensure_ascii=False, indent=2)
            output_text.delete("1.0", tk.END)
            output_text.insert(tk.END, json_output)
            status_label.config(text=f"解析完成，生成 {len(all_results)} 份结果")
        else:
            output_text.delete("1.0", tk.END)
            output_text.insert(tk.END, "未获取到有效数据")
            status_label.config(text="未获取到有效数据")

    select_button = ttk.Button(
        button_bar,
        text="选择周报文件",
        command=select_files,
        style="Accent.TButton",
    )
    select_button.pack(side=tk.LEFT, padx=(0, 10))

    def clear_output():
        output_text.delete("1.0", tk.END)
        status_label.config(text="已清空输出")

    clear_button = ttk.Button(button_bar, text="清空输出", command=clear_output)
    clear_button.pack(side=tk.LEFT)

    root.mainloop()


if __name__ == "__main__":
    main()
